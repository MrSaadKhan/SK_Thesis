import openpyxl
import os

def build_results(device_indices, dev1, dev2, mu_diff_device_seen, sigma_diff_device_seen, mu_diff_device_unseen, sigma_diff_device_unseen, file_name):
    # Create or load Excel file
    print('Updating result')

    excel_file_path = file_name + "_results.xlsx"
    if os.path.exists(excel_file_path):
        # Load existing Excel file
        book = openpyxl.load_workbook(excel_file_path)
    else:
        # Create new Excel file
        book = openpyxl.Workbook()

    # Convert device indices dictionary to DataFrame for indexing
    device_indices_df = list(device_indices.keys())

    # Get or create worksheet for each type of result
    for table_name in ['mu_diff_device_seen', 'sigma_diff_device_seen', 'mu_diff_device_unseen', 'sigma_diff_device_unseen']:
        if table_name in book.sheetnames:
            sheet = book[table_name]
        else:
            sheet = book.create_sheet(table_name)

            # Write device names in the first row and first column
            sheet.cell(row=1, column=1, value="Device")
            for i, device_name in enumerate(device_indices_df):
                cell = sheet.cell(row=1, column=i+2, value=device_name)
                cell.alignment = openpyxl.styles.Alignment(textRotation=90)  # Rotate text by 90 degrees
                sheet.cell(row=i+2, column=1, value=device_name)
                sheet.column_dimensions['A'].auto_size = True


    # Get device indices for dev1 and dev2
    dev1_index = device_indices_df.index(dev1) + 2
    dev2_index = device_indices_df.index(dev2) + 2

    # Update tables with results
    tables = {
        'mu_diff_device_seen': mu_diff_device_seen,
        'sigma_diff_device_seen': sigma_diff_device_seen,
        'mu_diff_device_unseen': mu_diff_device_unseen,
        'sigma_diff_device_unseen': sigma_diff_device_unseen
    }

    for table_name, data in tables.items():
        sheet = book[table_name]
        # Write results to the correct cell based on the order of dev1 and dev2
        sheet.cell(row=dev1_index, column=dev2_index, value=data)

    # Save Excel file
    book.save(excel_file_path)

    print('\033[92mResult Updated ✔\033[0m')
